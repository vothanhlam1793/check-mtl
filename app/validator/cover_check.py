import re
from datetime import datetime, timedelta
from typing import List, Optional, Tuple

import openpyxl

from config import EXCEL_EPOCH
from app.schemas import ErrorItem


def _search_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    pattern: str,
    max_row: int = 60,
    max_col: int = 50
) -> List[Tuple[int, int, str]]:
    matches = []
    pl = pattern.lower()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and pl in str(v).lower():
                matches.append((r, c, str(v)))
    return matches


def _get_neighbor(
    ws, row: int, col: int, dr: int = 0, dc: int = 1, max_r: int = 60, max_c: int = 50
) -> Optional[str]:
    nr, nc = row + dr, col + dc
    if 1 <= nr <= max_r and 1 <= nc <= max_c:
        v = ws.cell(row=nr, column=nc).value
        if v is not None:
            return str(v).strip()
    return None


def _extract_date(value, cell_ref: str = "") -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        if value > 30000:
            return EXCEL_EPOCH + timedelta(days=int(value))
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def check_cover(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    file_name: str = ""
) -> List[ErrorItem]:
    errors: List[ErrorItem] = []

    max_r = min(ws.max_row, 60)
    max_c = min(ws.max_column, 50)

    # ─── Rule 1: Tên công ty ───
    company_found = False
    company_candidates = _search_cells(ws, "công ty", max_r, max_c)
    tnhh_candidates = _search_cells(ws, "tnhh", max_r, max_c)
    all_comp = company_candidates + tnhh_candidates

    for row, col, val in all_comp:
        for dc in range(0, 8):
            neighbor = _get_neighbor(ws, row, col + dc, dc=0, max_r=max_r, max_c=max_c)
            if neighbor:
                nl = neighbor.lower()
                if "công ty" in nl or "tnhh" in nl:
                    if len(neighbor) > 10:
                        company_found = True
                        break
        if company_found:
            break

    if not company_found and tnhh_candidates:
        company_found = True

    if not company_found:
        errors.append(ErrorItem(
            row=0, col="Cover", field="Tên công ty",
            received="(trống)",
            reason="Cover thiếu tên công ty",
            severity="CRITICAL",
            fix="Điền tên công ty vào Cover (VD: 'Công ty TNHH...')"
        ))

    # ─── Rule 2: Quy mô dự án ───
    scale_label = _search_cells(ws, "quy mô", max_r, max_c)
    scale_filled = False
    for row, col, val in scale_label:
        for dc in range(1, 25):
            for dr in range(0, 2):
                n = _get_neighbor(ws, row, col + dc, dr=dr, max_r=max_r, max_c=max_c)
                if n and not n.startswith("=") and "quy mô" not in n.lower() and n not in (":", "-", ""):
                    if re.search(r'\d', n) or "ha" in n.lower() or "m2" in n.lower():
                        scale_filled = True
                        break
                    if len(n) > 2:
                        scale_filled = True
                        break
            if scale_filled:
                break
        if scale_filled:
            break

    if scale_label and not scale_filled:
        errors.append(ErrorItem(
            row=0, col="Cover", field="Quy mô dự án",
            received="(trống)",
            reason="Cover thiếu quy mô dự án",
            severity="WARNING",
            fix="Điền quy mô dự án (VD: '38.1 Ha')"
        ))

    # ─── Rule 3: Yêu cầu chung ───
    yc_label = _search_cells(ws, "yêu cầu chung", max_r, max_c)
    yc_filled = False
    for row, col, val in yc_label:
        for dc in range(1, 25):
            for dr in range(0, 2):
                n = _get_neighbor(ws, row, col + dc, dr=dr, max_r=max_r, max_c=max_c)
                if n and not n.startswith("=") and "yêu cầu" not in n.lower() and len(n) > 2:
                    yc_filled = True
                    break
            if yc_filled:
                break
        if yc_filled:
            break

    if yc_label and not yc_filled:
        errors.append(ErrorItem(
            row=0, col="Cover", field="Yêu cầu chung",
            received="(trống)",
            reason="Cover thiếu yêu cầu chung",
            severity="WARNING",
            fix="Điền yêu cầu chung của dự án"
        ))

    # ─── Rule 4: Rev > 1 + Điều chỉnh số trống ───
    rev_value: Optional[int] = None
    rev_label = _search_cells(ws, "lần ban hành", max_r, max_c)

    for row, col, val in rev_label:
        for dc in range(1, 10):
            for dr in range(0, 2):
                n = _get_neighbor(ws, row, col + dc, dr=dr, max_r=max_r, max_c=max_c)
                if n:
                    m = re.search(r'lần\s*(\d+)', n.lower())
                    if m:
                        rev_value = int(m.group(1))
                        break
                    try:
                        rev_value = int(n)
                        break
                    except (ValueError, TypeError):
                        pass
            if rev_value is not None:
                break
        if rev_value is not None:
            break

    if not rev_label and rev_value is None:
        rev_label2 = _search_cells(ws, "lần ban", max_r, max_c)
        for row, col, val in rev_label2:
            for dc in range(1, 10):
                for dr in range(0, 2):
                    n = _get_neighbor(ws, row, col + dc, dr=dr, max_r=max_r, max_c=max_c)
                    if n:
                        m = re.search(r'lần\s*(\d+)', n.lower())
                        if m:
                            rev_value = int(m.group(1))
                            break
                        try:
                            rev_value = int(n)
                            break
                        except (ValueError, TypeError):
                            pass
                if rev_value is not None:
                    break
            if rev_value is not None:
                break

    if rev_value and rev_value > 1:
        dc_label = _search_cells(ws, "điều chỉnh số", max_r, max_c)
        dc_filled = False
        for row, col, val in dc_label:
            for dc_off in range(1, 8):
                for dr_off in range(0, 3):
                    n = _get_neighbor(ws, row, col + dc_off, dr=dr_off, max_r=max_r, max_c=max_c)
                    if n and not n.startswith("=") and "điều chỉnh" not in n.lower():
                        try:
                            if int(n) >= 1:
                                dc_filled = True
                                break
                        except (ValueError, TypeError):
                            if len(n) > 0:
                                dc_filled = True
                                break
                if dc_filled:
                    break
            if dc_filled:
                break

        if not dc_filled:
            errors.append(ErrorItem(
                row=0, col="Cover", field="Điều chỉnh số",
                received="(trống)",
                reason=f"Lần ban hành > 1 (Rev={rev_value}) nhưng Điều chỉnh số trống",
                severity="WARNING",
                fix=f"Điền số Điều chỉnh tương ứng với Lần ban hành {rev_value}"
            ))

    # ─── Rule 5: Ngày lập > Ngày trong tên file ───
    file_date: Optional[datetime] = None
    file_date_match = re.search(r'(\d{8})', file_name)
    if file_date_match:
        try:
            file_date = datetime.strptime(file_date_match.group(1), "%Y%m%d")
        except ValueError:
            pass

    if file_date:
        date_keywords = ["ngà", "date", "ngay"]
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str):
                    sv = v.lower()
                    if any(kw in sv for kw in date_keywords):
                        for dr in range(-2, 3):
                            for dc_val in range(-1, 6):
                                if dr == 0 and dc_val == 0:
                                    continue
                                nr, nc = r + dr, c + dc_val
                                if 1 <= nr <= max_r and 1 <= nc <= max_c:
                                    nv = ws.cell(row=nr, column=nc).value
                                    dt = _extract_date(nv)
                                    if dt and dt > file_date:
                                        errors.append(ErrorItem(
                                            row=0, col="Cover",
                                            field="Ngày trên Cover",
                                            received=dt.strftime("%Y-%m-%d"),
                                            reason=(
                                                f"Ngày trên Cover ({dt.strftime('%Y-%m-%d')}) "
                                                f"sau ngày trong tên file ({file_date.strftime('%Y-%m-%d')})"
                                            ),
                                            severity="WARNING",
                                            fix="Kiểm tra lại ngày trên Cover"
                                        ))
                                        break
                            if errors:
                                break
                    if errors:
                        break
            if errors:
                break

    return errors
