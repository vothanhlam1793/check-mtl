import re
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Optional

import openpyxl

from config import DURATION_PATTERNS, KNOWN_STATUS_LIKE, WBS_PATTERN


@dataclass
class InferredLayout:
    header_row: int
    wbs_col: int = 0           # A = 0
    task_col: int = 1          # B = 1
    duration_col: int = 2      # C = 2 (or None if not found)
    start_col: int = 3         # D = 3 (or None)
    finish_col: int = 4        # E = 4 (or None)
    status_col: int = 5        # F or G (or None)
    notes_col: int = 6         # G (or None)
    extra_cols: List[int] = field(default_factory=list)
    has_extra_middle_col: bool = False  # e.g. Project Master Code at C
    col_offset: int = 0        # for EN_GMS where extra col at C shifts everything
    duration_pattern_index: int = 0    # which pattern matched
    sample_statuses: List[str] = field(default_factory=list)
    data_sheet_name: str = ""


def _is_date_value(v) -> bool:
    if v is None:
        return False
    if isinstance(v, datetime):
        return True
    if isinstance(v, (int, float)):
        # Excel serial dates: real project dates are 30000-80000 (2020s-2050s)
        # Real project dates are 30000+ (2020s-2050s)
        # Sentinel dates are 10000-20000 (e.g. 10959, 11403)
        # Values < 10000 are NOT dates (WBS codes, durations, etc.)
        return v > 30000 or (10000 < v < 20000)
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                datetime.strptime(v.strip(), fmt)
                return True
            except ValueError:
                continue
    return False


def _is_duration_value(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    for pat in DURATION_PATTERNS:
        if pat.match(s):
            return True
    return False


def _is_status_like(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    if len(s) > 40:
        return False
    return s in KNOWN_STATUS_LIKE


def infer_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> Optional[InferredLayout]:
    """Auto-detect column layout by position and data pattern.
    No template name matching needed.
    """

    # Step 1: Find header row — scan row 1-10 for "WBS" in column A
    header_row = None
    for r in range(1, 11):
        a_val = str(ws.cell(row=r, column=1).value or "").strip()
        if a_val.upper() == "WBS":
            header_row = r
            break

    if header_row is None:
        return None

    # Step 2: Sample data rows (first 15 after header)
    sample_rows = []
    for r in range(header_row + 1, min(header_row + 16, ws.max_row + 1)):
        row_vals = []
        for c in range(1, min(11, ws.max_column + 1)):
            row_vals.append(ws.cell(row=r, column=c).value)
        if any(v is not None and str(v).strip() for v in row_vals):
            sample_rows.append(row_vals)

    if not sample_rows:
        return None

    layout = InferredLayout(header_row=header_row)

    # Step 3: Infer each column by position + pattern
    # Column A (index 0) = always WBS
    layout.wbs_col = 0

    # Collect all values per column
    def col_values(col_idx):
        vals = []
        for row in sample_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                vals.append(row[col_idx])
        return vals

    # Step 4: For columns after A (B=1, C=2, ... G=6), infer role
    # Strategy: sample each column, classify by dominant pattern

    col_roles = {}  # col_idx -> role

    # First pass: classify obvious columns
    for ci in range(1, min(7, ws.max_column)):
        vals = col_values(ci)
        if not vals:
            continue

        # Count classification hits
        date_count = sum(1 for v in vals if _is_date_value(v))
        dur_count = sum(1 for v in vals if _is_duration_value(v))
        status_count = sum(1 for v in vals if _is_status_like(v))
        text_count = sum(1 for v in vals if isinstance(v, str) and len(str(v).strip()) > 0)
        formula_count = sum(1 for v in vals if isinstance(v, str) and v.startswith("="))
        total = len(vals)

        if total == 0:
            continue

        date_ratio = date_count / total
        dur_ratio = dur_count / total
        status_ratio = status_count / total
        text_ratio = text_count / total

        # Check if this is a middle extra column (like Project Master Code)
        # Indicators: text dominant, NOT date/duration/status, and between WBS and typical data columns
        if text_ratio > 0.6 and dur_ratio < 0.1 and date_ratio < 0.1 and status_ratio < 0.2:
            # Could be an extra column like Project Master Code or Task Name variant
            # Task name tends to have longer text
            avg_len = sum(len(str(v).strip()) for v in vals if isinstance(v, str)) / max(1, text_count)
            if avg_len > 20:
                col_roles[ci] = "task"
            else:
                # Short uniform text → likely project code or extra metadata
                col_roles[ci] = "extra_meta"

    # Check if this is a middle extra column (like Project Master Code)
    # Indicators: text dominant, NOT date/duration/status, and between WBS and typical data columns
    if text_ratio > 0.6 and dur_ratio < 0.1 and date_ratio < 0.1 and status_ratio < 0.2:
        # Could be an extra column like Project Master Code or Task Name variant
        avg_len = sum(len(str(v).strip()) for v in vals if isinstance(v, str)) / max(1, text_count)
        if avg_len > 20:
            col_roles[ci] = "task"
        else:
            # Short uniform text -> likely project code or extra metadata
            col_roles[ci] = "extra_meta"

    elif date_ratio > 0.3:
        col_roles[ci] = "date"
    elif dur_ratio > 0.3:
        # Double-check: if "duration" matches look like WBS codes (including numeric), it's extra_meta
        wbs_like = sum(1 for v in vals if v is not None and WBS_PATTERN.match(str(v).strip()))
        if wbs_like > len(vals) * 0.5 and ci < 4:
            col_roles[ci] = "extra_meta"
        else:
            col_roles[ci] = "duration"
    elif formula_count > 0.5 and ci == 2:
        col_roles[ci] = "duration"
    elif status_ratio > 0.3:
        col_roles[ci] = "status"

    # Step 5: Map roles to specific fields
    date_cols = sorted([ci for ci, role in col_roles.items() if role == "date"])
    task_cols = [ci for ci, role in col_roles.items() if role == "task"]
    dur_cols = [ci for ci, role in col_roles.items() if role == "duration"]
    status_cols = [ci for ci, role in col_roles.items() if role == "status"]
    extra_cols = [ci for ci, role in col_roles.items() if role == "extra_meta"]

    # Assign task column (pick first text-heavy column, usually B)
    if task_cols:
        layout.task_col = task_cols[0]
    else:
        layout.task_col = 1  # default B

    # Detect extra middle column BEFORE assignments (like Project Master Code at C in EN_GMS)
    if extra_cols and extra_cols[0] < 5:
        layout.has_extra_middle_col = True
        layout.col_offset = 1

    # Assign columns with offset awareness
    if layout.has_extra_middle_col and layout.col_offset > 0:
        offset = layout.col_offset
        layout.duration_col = 2 + offset
        layout.start_col = 3 + offset
        layout.finish_col = 4 + offset
        layout.status_col = 5 + offset
        layout.notes_col = 6 + offset
    else:
        # Assign duration column
        if dur_cols:
            layout.duration_col = dur_cols[0]
        else:
            layout.duration_col = 2

        # Assign status column: prefer G(6) if it has short text, else F(5)
        g_text = [v for v in col_values(6) if isinstance(v, str) and 1 <= len(str(v).strip()) <= 40]
        f_text = [v for v in col_values(5) if isinstance(v, str) and 1 <= len(str(v).strip()) <= 40]
        if len(g_text) >= len(f_text):
            layout.status_col = 6
        elif status_cols:
            layout.status_col = status_cols[0]
        elif f_text:
            layout.status_col = 5
        else:
            layout.status_col = 6

        # Notes column = last unassigned within range
        assigned = {layout.wbs_col, layout.task_col, layout.duration_col,
                    layout.start_col, layout.finish_col, layout.status_col}
        for ci in range(6, 0, -1):
            if ci not in assigned:
                layout.notes_col = ci
                break
        else:
            layout.notes_col = 6

    # Collect sample statuses for validation
    status_vals = col_values(layout.status_col)
    layout.sample_statuses = sorted(set(
        str(v).strip() for v in status_vals if v and isinstance(v, str) and len(str(v).strip()) < 40
    ))[:10]

    return layout
