import openpyxl
from typing import Tuple, Optional

from config import COVER_SHEET_NAMES, TEMPLATE_SHEET_NAMES, ALL_TEMPLATES, TemplateConfig


class FileCheckError(Exception):
    pass


def check_and_open_file(file_path: str) -> Tuple[openpyxl.Workbook, str, str, TemplateConfig, int]:
    """Step A: open file, verify sheets, detect template format and header row.
    Returns (workbook, data_sheet_name, cover_sheet_name, template_config, header_row).
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise FileCheckError(f"Không thể mở file. Lỗi: {e}")

    sheet_names = wb.sheetnames

    cover_sheet = None
    data_sheet = None
    data_sheet_name = ""

    for name in sheet_names:
        if name in COVER_SHEET_NAMES:
            cover_sheet = name
        elif name in TEMPLATE_SHEET_NAMES:
            continue
        elif name.lower().startswith("cover"):
            cover_sheet = name
        else:
            data_sheet = name
            data_sheet_name = name

    if not cover_sheet:
        wb.close()
        raise FileCheckError("Thiếu sheet 'Cover'. Sheet hiện có: " + ", ".join(sheet_names))

    if not data_sheet:
        wb.close()
        raise FileCheckError("Không tìm thấy sheet dữ liệu dự án. Sheet hiện có: " + ", ".join(sheet_names))

    ws = wb[data_sheet_name]

    # Auto-detect template by scanning header rows
    template, header_row = _detect_template(ws)
    if not template:
        wb.close()
        raise FileCheckError("Không thể xác định template. Header không khớp với mẫu nào.")

    return wb, data_sheet_name, cover_sheet, template, header_row


def _detect_template(ws) -> Tuple[Optional[TemplateConfig], int]:
    """Scan rows 1-10 to find header row and match to a template."""
    for r in range(1, 11):
        actual_headers = {}
        for c in range(1, 8):
            val = str(ws.cell(row=r, column=c).value or "").strip()
            actual_headers[chr(64 + c)] = val

        for tpl in ALL_TEMPLATES:
            # Check if headers match at least 5/7 columns
            match_count = 0
            for col_letter, expected in tpl.columns.items():
                actual = actual_headers.get(col_letter, "")
                if actual == expected:
                    match_count += 1
            if match_count >= 5:
                return tpl, r

    return None, 0
